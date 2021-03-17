import collections

import neo4j
from django.shortcuts import render, redirect
from neo4j import GraphDatabase
from project.settings import neo4jUser, neo4jPass, neo4jURI

neo4jUsername = neo4jUser
neo4jPassword = neo4jPass
neo4jUri = neo4jURI

from SlaGenerator.models import MACM, Asset, Relation, Protocol, Attribute, Attribute_value, Asset_Attribute_value, \
    Threat_Attribute_value, Threat_CIA, Threat_Stride


def apps_management(request):
    ordered_apps = []
    context = {}
    try:
        graphDriver = GraphDatabase.driver(uri=neo4jUri, auth=(neo4jUsername, neo4jPassword))
        session = graphDriver.session()
        nodes_string = session.run("match (node) return node")
        nodes = [record for record in nodes_string.data()]
        apps = {}
        for node in nodes:
            try:
                apps[node['node']['app_id']] = node['node']['application']
            except:
                print("Cannot parse graph with node " + str(node['node']))
        ordered_apps = collections.OrderedDict(sorted(apps.items()))
        # print(ordered_apps)
        for appId, application in ordered_apps.items():
            MACM_instance = MACM(appId=appId, application=application)
            MACMvalue = MACM.objects.all().filter(appId=appId, application=application)
            if not MACMvalue:
                MACM_instance.save()
            graphDriver.close()
        context = {
            'apps': ordered_apps
        }
    except neo4j.exceptions.ServiceUnavailable as error:
        print(error)
        context = {
            'error': error
        }
    return render(request, 'apps_management.html', context)


def get_graphNodesbyAppId(appId):
    graph = GraphDatabase.driver(uri=neo4jUri, auth=(neo4jUsername, neo4jPassword))
    session = graph.session()
    nodes_string = session.run("MATCH (node { app_id:  \'" + str(appId) + "\' }) RETURN "
                                                                          "node,labels(node) as nodeType")
    nodes = [record for record in nodes_string.data()]
    session.close()
    return nodes


def get_graphRelationbyAppId(appId):
    graph = GraphDatabase.driver(uri=neo4jUri, auth=(neo4jUsername, neo4jPassword))
    session = graph.session()
    nodes_string = session.run("MATCH (client { app_id:  \'" + str(
        appId) + "\' }) -[relation]->(server) RETURN client,labels(client) as clientType,"
                 " relation,TYPE(relation) as relationType,relation.protocol as protocol, "
                 "server,labels(server) as serverType")
    nodes = [record for record in nodes_string.data()]
    session.close()
    return nodes


def macm_viewer(request, appId):
    return render(request, 'macm_viewer.html', {"appId": appId})


def asset_management(request, appId):
    # save assets info in sqlite
    # nodes = Asset.objects.all().filter(app=MACM.objects.get(appId=appId))
    # metto nodes=None perchè così prende sempre fa neo4j (dovrei gestire la coerenza fra i due DB)
    nodes = None
    # connect to neo4j only if sqlite assets are empty (API are laggy)
    if not nodes:
        nodes = get_graphNodesbyAppId(appId)
        for node in nodes:
            # print(node["node"]["name"]+" "+ node["node"]["type"])
            # print(node)

            asset = Asset.objects.all().get_or_create(app=MACM.objects.get(appId=appId),
                                                      name=node["node"]["name"])
            # mi salvo id sqlite in dizionario
            node['id'] = asset[0].id

            try:
                # vedo se il nome del componente è un attribute value
                # per il 5g andrebbero considerate le minacce sia di SERVICE.Web che di UE (ad esempio)
                Attribute_value_instance = Attribute_value.objects.get(attribute_value=node["node"]["type"])
                Asset_Attribute_value.objects.all().get_or_create(asset=asset[0],
                                                                  attribute_value=Attribute_value_instance)
                nodes = Asset_Attribute_value.objects.all().filter(app=MACM.objects.get(appId=appId))
            except:
                print()

    # save relation info in sqlite

    arches = get_graphRelationbyAppId(appId)
    for arch in arches:
        Asset_client = Asset.objects.all().filter(name=arch["client"]["name"], app=MACM.objects.get(appId=appId))
        Asset_server = Asset.objects.all().filter(name=arch["server"]["name"], app=MACM.objects.get(appId=appId))

        if arch["protocol"] is None:
            print()
        elif isinstance(arch["protocol"], str):
            # single protocol in one arch
            try:
                Relation.objects.all().get_or_create(asset=Asset.objects.get(name=arch["client"]["name"],
                                                                             app=MACM.objects.get(appId=appId)),
                                                     protocol=Protocol.objects.get(protocol=arch["protocol"]),
                                                     app=MACM.objects.get(appId=appId),
                                                     relation_type=arch["relationType"],
                                                     role="client")
                Relation.objects.all().get_or_create(asset=Asset.objects.get(name=arch["server"]["name"],
                                                                             app=MACM.objects.get(appId=appId)),
                                                     protocol=Protocol.objects.get(protocol=arch["protocol"]),
                                                     app=MACM.objects.get(appId=appId),
                                                     relation_type=arch["relationType"],
                                                     role="server")
            except:
                print()
                # print("Protocol info not found in arch between " + str(arch["client"]["name"]) + " and " + str(
                # arch["server"]["name"]))
        elif isinstance(arch["protocol"], list):
            for protocol in arch["protocol"]:
                # print(protocol)
                # multiple protocol in one arch
                try:
                    Relation.objects.all().get_or_create(asset=Asset.objects.get(name=arch["client"]["name"],
                                                                                 app=MACM.objects.get(appId=appId)),
                                                         protocol=Protocol.objects.get(protocol=protocol),
                                                         app=MACM.objects.get(appId=appId),
                                                         relation_type=arch["relationType"],
                                                         role="client")
                    Relation.objects.all().get_or_create(asset=Asset.objects.get(name=arch["server"]["name"],
                                                                                 app=MACM.objects.get(appId=appId)),
                                                         protocol=Protocol.objects.get(protocol=protocol),
                                                         app=MACM.objects.get(appId=appId),
                                                         relation_type=arch["relationType"],
                                                         role="server")
                except:
                    print()
                    # print("Protocol info not found in arch between " + str(arch["client"]["name"]) + " and " + str(
                    #   arch["server"]["name"]))
            else:
                print("error getting protocol information")
        # we consider only relations with some associated properties
        relations = Relation.objects.all().filter(app=MACM.objects.get(appId=appId))
    return render(request, 'asset_management.html', {
        'nodes': nodes,
        'relations': relations,
        'appId': appId
    })


def threat_modeling_per_asset(request, appId, assetId):
    threats = []
    asset = Asset.objects.all().filter(id=assetId)[0]
    asset_attribute_value = Asset_Attribute_value.objects.all().filter(asset_id=assetId)
    threats_attribute_values = Threat_Attribute_value.objects.all().filter(
        attribute_value_id=asset_attribute_value[0].attribute_value.id)
    for threat_attribute_value in threats_attribute_values:
        strides_per_threat = []
        affectedRequirements = []
        try:
            for stride in Threat_Stride.objects.all().filter(threat=threat_attribute_value.threat):
                strides_per_threat.append(stride.stride.category)
            for requirement in Threat_CIA.objects.all().filter(threat=threat_attribute_value.threat):
                affectedRequirements.append(requirement.cia.requirement)
        except:
            print("Error in selecting additional info")

        threats.append((threat_attribute_value.threat, strides_per_threat, affectedRequirements))

    return render(request, 'threat_modeling_per_asset.html', {
        'threats': threats,
        'asset': asset}
                  )


def threat_modeling(request, appId):
    threats_list = []
    nodes = get_graphNodesbyAppId(appId)
    for node in nodes:
        asset = Asset.objects.all().filter(name=node["node"]["name"])[0]
        asset_attribute_value = Asset_Attribute_value.objects.all().filter(asset_id=asset.id)
        try:
            print(asset.name + " " + asset_attribute_value[0].attribute_value.attribute_value)
            threats_attribute_values = Threat_Attribute_value.objects.all().filter(
                attribute_value_id=asset_attribute_value[0].attribute_value.id)
            for threat_attribute_value in threats_attribute_values:
                strides_per_threat = []
                affectedRequirements = []
                try:
                    for stride in Threat_Stride.objects.all().filter(threat=threat_attribute_value.threat):
                        strides_per_threat.append(stride.stride.category)
                    for requirement in Threat_CIA.objects.all().filter(threat=threat_attribute_value.threat):
                        affectedRequirements.append(requirement.cia.requirement)
                except:
                    print("Error in selecting additional info")

                threats_list.append(
                    (threat_attribute_value.threat, strides_per_threat, affectedRequirements, asset.name))

        except:
            print()

    return render(request, 'threat_modeling.html', {"threats": threats_list, "appId": appId
                                                    })


from openpyxl.styles import Font, Border, Side
import csv
from datetime import datetime

from django.http import HttpResponse
from django.shortcuts import render, redirect
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side


def export_threat_modeling(request, appId):
    if request.method == "POST":

        # help: https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-{name}-TM-report.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'),
            name=MACM.objects.get(appId=appId).application.replace(" ", "_")
        )
        workbook = Workbook()

        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Threat_modeling_REPORT'
        columns = ['#','Asset name', 'Asset type', 'Threat',  'CIA', 'STRIDE','Description']
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.font = Font(name="Times New Roman", size=12, bold=True, color='FF0000')
            cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                 right=Side(border_style="thin", color='FF000000'),
                                 top=Side(border_style="thin", color='FF000000'),
                                 bottom=Side(border_style="thin", color='FF000000'), )

        threats_list = []
        nodes = get_graphNodesbyAppId(appId)
        for node in nodes:
            asset = Asset.objects.all().filter(name=node["node"]["name"])[0]
            asset_attribute_value = Asset_Attribute_value.objects.all().filter(asset_id=asset.id)
            try:
                threats_attribute_values = Threat_Attribute_value.objects.all().filter(
                attribute_value_id=asset_attribute_value[0].attribute_value.id)
                for threat_attribute_value in threats_attribute_values:
                    strides_per_threat = []
                    affectedRequirements = []
                    try:
                        for stride in Threat_Stride.objects.all().filter(threat=threat_attribute_value.threat):
                            strides_per_threat.append(stride.stride.category)
                        for requirement in Threat_CIA.objects.all().filter(threat=threat_attribute_value.threat):
                            affectedRequirements.append(requirement.cia.requirement)
                    except:
                        print("Error in selecting additional info")

                    threats_list.append(
                        (threat_attribute_value.threat, strides_per_threat, affectedRequirements,
                         asset.name,threat_attribute_value.attribute_value))
            except:
                print()


        for threat in threats_list:
            print(threat)
            row_num += 1
            stride=""
            cia=""
            for index,single in enumerate(threat[1]):
                if not index==len(threat[1])-1:
                    stride+=single+", "
                else:
                    stride+=single

            for index,single in enumerate(threat[2]):
                if not index==len(threat[1])-1:
                    cia+=single+", "
                else:
                    cia+=single

            # columns = ['Asset name', 'Asset type', 'Threat', 'Description', 'CIA', 'STRIDE']

            # Define the data for each cell in the row
            row = [
                row_num,
                threat[3],
                threat[4].attribute_value,
                threat[0].name,
                cia,
                stride,
                threat[0].description,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                     right=Side(border_style="thin", color='FF000000'),
                                     top=Side(border_style="thin", color='FF000000'),
                                     bottom=Side(border_style="thin", color='FF000000'), )

                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value
                    cell.font = Font(name="Times New Roman", size=11, bold=False, color='FF000000')
                    cell.border = Border(left=Side(border_style="thin", color='FF000000'),
                                         right=Side(border_style="thin", color='FF000000'),
                                         top=Side(border_style="thin", color='FF000000'),
                                         bottom=Side(border_style="thin", color='FF000000'), )
        # Per effettuare il resize delle celle in base a quella più grande
        dims = {}

        from openpyxl.styles import Alignment

        for row in worksheet.rows:
            for cell in row:
                cell.alignment = Alignment(wrap_text=True)
                if cell.value:
                    dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))+0.05
        for col, value in dims.items():
            worksheet.column_dimensions[col].width = value


        workbook.save(response)

        return response
